"""
文档智能操作交互模块 - Excel 文档操作实现
支持：单元格编辑、行列操作、格式调整、数据提取等操作
"""
import os
import re
import json
from typing import Dict, List, Optional, Any, Tuple, Union
from loguru import logger
from openai import OpenAI
from config import settings

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────────────────
# Excel 文档操作类
# ─────────────────────────────────────────────────────────────────────────────
class ExcelDocumentOperations:
    """
    Excel 文档操作类
    提供基于自然语言指令的表格编辑功能
    """
    
    def __init__(self, file_path: str):
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl 未安装，请运行: pip install openpyxl")
        
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
    
    def _parse_cell_ref(self, ref: str) -> Tuple[int, int]:
        """
        解析单元格引用
        
        Args:
            ref: 单元格引用，如 "A1", "第2行第3列", "第3行第2列"
            
        Returns:
            (row, col) 1-indexed
        """
        # 直接的 A1 格式
        match = re.match(r'^([A-Z]+)(\d+)$', ref.upper())
        if match:
            col_str = match.group(1)
            row = int(match.group(2))
            # 转换列字母为数字
            col = 0
            for char in col_str:
                col = col * 26 + (ord(char) - ord('A') + 1)
            return row, col
        
        # 中文格式 "第X行第Y列"
        match = re.match(r'第(\d+)[行个]\s*第(\d+)[列表]', ref)
        if match:
            return int(match.group(1)), int(match.group(2))
        
        # 中文格式 "第Y列第X行"
        match = re.match(r'第(\d+)[列表]\s*第(\d+)[行个]', ref)
        if match:
            return int(match.group(2)), int(match.group(1))
        
        # 简化的 "X,Y" 格式
        match = re.match(r'^(\d+),(\d+)$', ref)
        if match:
            return int(match.group(1)), int(match.group(2))
        
        return 1, 1  # 默认返回 A1
    
    def _cell_ref_to_coords(self, row: int, col: int) -> str:
        """将行列号转换为单元格引用"""
        return f"{get_column_letter(col)}{row}"
    
    # ─────────────────────────────────────────────────────────────────────────
    # 基础操作：读取
    # ─────────────────────────────────────────────────────────────────────────
    
    def get_all_content(self) -> Dict[str, Any]:
        """获取工作表所有内容"""
        sheets = {}
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            data = []
            
            for row in ws.iter_rows(min_row=1, max_row=min(100, ws.max_row), values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(row_data):  # 只保留非空行
                    data.append(row_data)
            
            sheets[sheet_name] = {
                'rows': ws.max_row,
                'cols': ws.max_column,
                'data': data[:50]  # 预览前50行
            }
        
        return {
            'success': True,
            'sheets': sheets,
            'active_sheet': self.ws.title
        }
    
    def get_cell_value(self, cell_ref: str) -> Any:
        """获取单元格值"""
        row, col = self._parse_cell_ref(cell_ref)
        return self.ws.cell(row=row, column=col).value
    
    def get_row_values(self, row_num: int) -> List[Any]:
        """获取整行数据"""
        return [cell.value for cell in self.ws[row_num]]
    
    def get_column_values(self, col_num: int) -> List[Any]:
        """获取整列数据"""
        return [self.ws.cell(row=r, column=col_num).value for r in range(1, self.ws.max_row + 1)]
    
    def find_cells(self, keyword: str, sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        搜索包含关键词的单元格
        
        Args:
            keyword: 搜索关键词
            sheet_name: 工作表名称（可选）
            
        Returns:
            匹配的单元格列表
        """
        results = []
        search_sheets = [self.wb[sheet_name]] if sheet_name else self.wb.worksheets
        
        for ws in search_sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and keyword in str(cell.value):
                        results.append({
                            'sheet': ws.title,
                            'cell': self._cell_ref_to_coords(cell.row, cell.column),
                            'value': str(cell.value)
                        })
        
        return results
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：编辑单元格
    # ─────────────────────────────────────────────────────────────────────────
    
    def edit_cell(self, cell_ref: str, value: Any, 
                  preserve_format: bool = True) -> Dict[str, Any]:
        """
        编辑单元格
        
        Args:
            cell_ref: 单元格引用
            value: 新值
            preserve_format: 是否保留原有格式
            
        Returns:
            操作结果
        """
        try:
            row, col = self._parse_cell_ref(cell_ref)
            old_value = self.ws.cell(row=row, column=col).value
            
            self.ws.cell(row=row, column=col, value=value)
            self.wb.save(self.file_path)
            
            return {
                'success': True,
                'message': f'已将 {cell_ref} 从 "{old_value}" 改为 "{value}"',
                'old_value': old_value,
                'new_value': value,
                'cell': cell_ref
            }
        except Exception as e:
            return {'success': False, 'message': f'编辑失败: {str(e)}'}
    
    def batch_edit_cells(self, updates: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        批量编辑单元格
        
        Args:
            updates: 更新列表 [{"cell": "A1", "value": "xxx"}, ...]
            
        Returns:
            操作结果
        """
        success_count = 0
        failed = []
        
        for update in updates:
            cell_ref = update.get('cell', '')
            value = update.get('value')
            
            try:
                row, col = self._parse_cell_ref(cell_ref)
                self.ws.cell(row=row, column=col, value=value)
                success_count += 1
            except Exception as e:
                failed.append({'cell': cell_ref, 'error': str(e)})
        
        if success_count > 0:
            self.wb.save(self.file_path)
        
        return {
            'success': len(failed) == 0,
            'message': f'成功更新 {success_count} 个单元格' + (f'，失败 {len(failed)} 个' if failed else ''),
            'success_count': success_count,
            'failed': failed
        }
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：行列操作
    # ─────────────────────────────────────────────────────────────────────────
    
    def add_row(self, position: int = None, values: List[Any] = None) -> Dict[str, Any]:
        """
        添加行
        
        Args:
            position: 添加位置（行号），None表示末尾
            values: 该行的数据
            
        Returns:
            操作结果
        """
        try:
            if position is None:
                position = self.ws.max_row + 1
            
            self.ws.insert_rows(position)
            
            if values:
                for col, val in enumerate(values, start=1):
                    self.ws.cell(row=position, column=col, value=val)
            
            self.wb.save(self.file_path)
            
            return {
                'success': True,
                'message': f'已在第 {position} 行插入新行',
                'row': position
            }
        except Exception as e:
            return {'success': False, 'message': f'插入行失败: {str(e)}'}
    
    def add_column(self, position: int = None, header: str = None) -> Dict[str, Any]:
        """
        添加列
        
        Args:
            position: 添加位置（列号），None表示末尾
            header: 列标题
            
        Returns:
            操作结果
        """
        try:
            if position is None:
                position = self.ws.max_column + 1
            
            self.ws.insert_cols(position)
            
            if header:
                self.ws.cell(row=1, column=position, value=header)
            
            self.wb.save(self.file_path)
            
            col_letter = get_column_letter(position)
            return {
                'success': True,
                'message': f'已在第 {position} 列插入新列（{col_letter}列）',
                'column': position,
                'column_letter': col_letter
            }
        except Exception as e:
            return {'success': False, 'message': f'插入列失败: {str(e)}'}
    
    def delete_row(self, row_num: int) -> Dict[str, Any]:
        """删除行"""
        try:
            self.ws.delete_rows(row_num)
            self.wb.save(self.file_path)
            return {
                'success': True,
                'message': f'已删除第 {row_num} 行',
                'row': row_num
            }
        except Exception as e:
            return {'success': False, 'message': f'删除行失败: {str(e)}'}
    
    def delete_column(self, col_num: int) -> Dict[str, Any]:
        """删除列"""
        try:
            col_letter = get_column_letter(col_num)
            self.ws.delete_cols(col_num)
            self.wb.save(self.file_path)
            return {
                'success': True,
                'message': f'已删除第 {col_num} 列（{col_letter}列）',
                'column': col_num
            }
        except Exception as e:
            return {'success': False, 'message': f'删除列失败: {str(e)}'}
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：格式调整
    # ─────────────────────────────────────────────────────────────────────────
    
    def format_cell(self, cell_ref: str, **format_params) -> Dict[str, Any]:
        """
        格式化单元格
        
        Args:
            cell_ref: 单元格引用
            **format_params: 格式参数
                - bold: 加粗
                - italic: 倾斜
                - font_size: 字号
                - font_color: 字体颜色
                - bg_color: 背景颜色
                - alignment: 对齐方式 (left, center, right)
                
        Returns:
            操作结果
        """
        try:
            row, col = self._parse_cell_ref(cell_ref)
            cell = self.ws.cell(row=row, column=col)
            changes = []
            
            # 字体加粗
            if format_params.get('bold'):
                if cell.font:
                    cell.font = Font(bold=True, name=cell.font.name, size=cell.font.size)
                else:
                    cell.font = Font(bold=True)
                changes.append('加粗')
            
            # 字体倾斜
            if format_params.get('italic'):
                if cell.font:
                    cell.font = Font(italic=True, name=cell.font.name, size=cell.font.size)
                else:
                    cell.font = Font(italic=True)
                changes.append('倾斜')
            
            # 字号
            font_size = format_params.get('font_size')
            if font_size:
                if cell.font:
                    cell.font = Font(size=font_size, name=cell.font.name,
                                   bold=cell.font.bold, italic=cell.font.italic)
                else:
                    cell.font = Font(size=font_size)
                changes.append(f'字号{font_size}')
            
            # 字体颜色
            font_color = format_params.get('font_color')
            if font_color:
                if cell.font:
                    cell.font = Font(color=font_color, name=cell.font.name,
                                   size=cell.font.size, bold=cell.font.bold)
                else:
                    cell.font = Font(color=font_color)
                changes.append(f'颜色{font_color}')
            
            # 背景颜色
            bg_color = format_params.get('bg_color')
            if bg_color:
                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                changes.append(f'背景{font_color}')
            
            # 对齐
            alignment = format_params.get('alignment')
            if alignment:
                align_map = {
                    'left': 'left', 'center': 'center', 'right': 'right',
                    '两端对齐': 'justify'
                }
                cell.alignment = Alignment(horizontal=align_map.get(alignment, 'left'))
                changes.append(f'对齐{alignment}')
            
            self.wb.save(self.file_path)
            
            return {
                'success': True,
                'message': f'已设置 {cell_ref}: {", ".join(changes) if changes else "无变化"}',
                'changes': changes
            }
        except Exception as e:
            return {'success': False, 'message': f'格式化失败: {str(e)}'}
    
    def format_row(self, row_num: int, **format_params) -> Dict[str, Any]:
        """格式化整行"""
        try:
            changes = []
            for col in range(1, self.ws.max_column + 1):
                result = self.format_cell(self._cell_ref_to_coords(row_num, col), **format_params)
                if result.get('success') and result.get('changes'):
                    changes.extend(result.get('changes', []))
            
            return {
                'success': True,
                'message': f'已格式化第 {row_num} 行',
                'changes': list(set(changes))
            }
        except Exception as e:
            return {'success': False, 'message': f'格式化行失败: {str(e)}'}
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：数据提取
    # ─────────────────────────────────────────────────────────────────────────
    
    def extract_table(self, sheet_name: str = None, 
                      has_header: bool = True) -> Dict[str, Any]:
        """
        提取表格数据
        
        Args:
            sheet_name: 工作表名称，None表示当前活动表
            has_header: 是否有表头
            
        Returns:
            表格数据
        """
        try:
            ws = self.wb[sheet_name] if sheet_name else self.ws
            
            headers = None
            rows = []
            
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if has_header and i == 1:
                    headers = [str(h) if h else f'列{col}' for col, h in enumerate(row, start=1)]
                else:
                    rows.append([cell for cell in row])
            
            return {
                'success': True,
                'sheet': ws.title,
                'headers': headers,
                'rows': rows,
                'row_count': len(rows),
                'col_count': len(headers) if headers else 0
            }
        except Exception as e:
            return {'success': False, 'message': f'提取表格失败: {str(e)}'}
    
    def extract_column(self, col_identifier: Union[int, str]) -> Dict[str, Any]:
        """
        提取整列数据
        
        Args:
            col_identifier: 列号或列标题
            
        Returns:
            列数据
        """
        try:
            # 如果是标题字符串，找到对应的列号
            if isinstance(col_identifier, str):
                found_col = None
                for row in self.ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        if str(cell.value) == col_identifier:
                            found_col = cell.column
                            break
                if found_col is None:
                    return {'success': False, 'message': f'未找到列标题: {col_identifier}'}
                col_num = found_col
            else:
                col_num = col_identifier
            
            # 提取数据
            values = []
            for row in self.ws.iter_rows(min_row=2, min_col=col_num, max_col=col_num, values_only=True):
                if row[0] is not None:
                    values.append(row[0])
            
            return {
                'success': True,
                'column': col_num,
                'values': values,
                'count': len(values)
            }
        except Exception as e:
            return {'success': False, 'message': f'提取列失败: {str(e)}'}
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：计算
    # ─────────────────────────────────────────────────────────────────────────
    
    def calculate(self, formula: str, cell_ref: str = None) -> Dict[str, Any]:
        """
        执行计算
        
        Args:
            formula: 计算公式（带=号），如 "=SUM(A1:A10)"
            cell_ref: 结果写入的单元格
            
        Returns:
            计算结果
        """
        try:
            if not formula.startswith('='):
                formula = '=' + formula
            
            if cell_ref:
                row, col = self._parse_cell_ref(cell_ref)
                self.ws.cell(row=row, column=col, value=formula)
                self.wb.save(self.file_path)
                return {
                    'success': True,
                    'message': f'已在 {cell_ref} 写入公式: {formula}',
                    'formula': formula
                }
            else:
                # 只计算不写入
                result = self.wb['Sheet1'].parent.calculate
                return {
                    'success': True,
                    'formula': formula,
                    'note': '请指定写入单元格'
                }
        except Exception as e:
            return {'success': False, 'message': f'计算失败: {str(e)}'}
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：批量填充（智能填表核心）
    # ─────────────────────────────────────────────────────────────────────────
    
    def batch_fill(self, field_mapping: Dict[str, List[Any]]) -> Dict[str, Any]:
        """
        批量填充数据（根据表头自动匹配列）
        
        Args:
            field_mapping: {字段名: [值列表]}
            
        Returns:
            操作结果
        """
        try:
            # 找到表头行
            header_row = 1
            header_map = {}  # {列号: 字段名}
            
            for row in self.ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value:
                        header_map[cell.column] = str(cell.value)
            
            # 匹配并填充
            filled = 0
            matched_fields = []
            
            for col_num, header in header_map.items():
                # 模糊匹配字段名
                for field_name, values in field_mapping.items():
                    if (field_name.lower() in header.lower() or 
                        header.lower() in field_name.lower()):
                        
                        # 填充数据
                        for row_offset, value in enumerate(values):
                            target_row = header_row + 1 + row_offset
                            self.ws.cell(row=target_row, column=col_num, value=value)
                            filled += 1
                        
                        matched_fields.append(header)
                        break
            
            self.wb.save(self.file_path)
            
            return {
                'success': True,
                'message': f'成功填充 {filled} 个单元格',
                'matched_fields': matched_fields,
                'filled_count': filled
            }
        except Exception as e:
            return {'success': False, 'message': f'批量填充失败: {str(e)}'}
    
    # ─────────────────────────────────────────────────────────────────────────
    # 操作：条件操作（条件格式 / 条件删除 / 条件筛选）
    # ─────────────────────────────────────────────────────────────────────────

    def _find_column_index(self, column: str) -> Optional[int]:
        """
        根据列名（表头文字）或列号字符串找到列索引（1-based）。
        返回 None 表示未找到。
        """
        if not column:
            return None
        # 纯数字：直接当列号
        if str(column).isdigit():
            return int(column)
        # 单个大写字母：A→1, B→2 ...
        col_upper = str(column).strip().upper()
        if re.match(r'^[A-Z]+$', col_upper):
            idx = 0
            for ch in col_upper:
                idx = idx * 26 + (ord(ch) - ord('A') + 1)
            return idx
        # 按表头文字模糊匹配（第1行）
        for cell in self.ws[1]:
            if cell.value and str(column).strip() in str(cell.value):
                return cell.column
        return None

    def _eval_condition(self, value, condition: str) -> bool:
        """
        判断单元格值是否满足条件字符串。
        condition 格式：">90" / ">=90" / "<60" / "==0" / "!=0"
                        "contains:关键词" / "empty" / "not_empty"
        """
        if condition == "empty":
            return value is None or str(value).strip() == ""
        if condition == "not_empty":
            return value is not None and str(value).strip() != ""
        if condition.startswith("contains:"):
            keyword = condition[len("contains:"):]
            return keyword in str(value) if value is not None else False

        # 数值比较
        try:
            num = float(value)
        except (TypeError, ValueError):
            return False

        m = re.match(r'^(>=|<=|!=|>|<|==)\s*(-?[\d.]+)$', condition.strip())
        if not m:
            return False
        op, threshold = m.group(1), float(m.group(2))
        return {
            '>': num > threshold, '>=': num >= threshold,
            '<': num < threshold, '<=': num <= threshold,
            '==': num == threshold, '!=': num != threshold,
        }.get(op, False)

    def conditional_format(self, column: str, condition: str, color: str,
                           sheet_name: str = None) -> Dict[str, Any]:
        """
        对满足条件的单元格批量设置字体颜色。

        Args:
            column:    列名（表头文字）或列号（"A"/"1"/数字），None 表示扫描所有列
            condition: 条件字符串，如 ">90"、"contains:优秀"、"empty"
            color:     十六进制颜色，如 "FF0000"
            sheet_name: 工作表名，None 表示当前活动表
        Returns:
            操作结果，包含受影响的单元格列表
        """
        try:
            ws = self.wb[sheet_name] if sheet_name else self.ws
            col_idx = self._find_column_index(column) if column else None

            affected = []
            # 数据从第2行开始（跳过表头）
            for row in ws.iter_rows(min_row=2):
                cells_to_check = [row[col_idx - 1]] if col_idx else list(row)
                for cell in cells_to_check:
                    if self._eval_condition(cell.value, condition):
                        existing = cell.font
                        cell.font = Font(
                            color=color,
                            bold=existing.bold if existing else False,
                            italic=existing.italic if existing else False,
                            size=existing.size if existing else None,
                            name=existing.name if existing else None,
                        )
                        affected.append(f"{cell.coordinate}({cell.value})")

            self.wb.save(self.file_path)
            return {
                'success': True,
                'message': f'条件格式完成，共影响 {len(affected)} 个单元格',
                'affected_cells': affected,
                'affected_count': len(affected),
            }
        except Exception as e:
            return {'success': False, 'message': f'条件格式失败: {str(e)}'}

    def conditional_delete(self, column: str, condition: str,
                           sheet_name: str = None) -> Dict[str, Any]:
        """
        删除满足条件的整行。

        Args:
            column:    列名或列号，None 表示只要该行任意单元格满足条件即删除
            condition: 条件字符串
            sheet_name: 工作表名
        Returns:
            操作结果
        """
        try:
            ws = self.wb[sheet_name] if sheet_name else self.ws
            col_idx = self._find_column_index(column) if column else None

            # 从后往前删，避免行号偏移
            rows_to_delete = []
            for row in ws.iter_rows(min_row=2):
                cells_to_check = [row[col_idx - 1]] if col_idx else list(row)
                if any(self._eval_condition(c.value, condition) for c in cells_to_check):
                    rows_to_delete.append(row[0].row)

            for row_num in reversed(rows_to_delete):
                ws.delete_rows(row_num)

            self.wb.save(self.file_path)
            return {
                'success': True,
                'message': f'条件删除完成，共删除 {len(rows_to_delete)} 行',
                'deleted_rows': rows_to_delete,
                'deleted_count': len(rows_to_delete),
            }
        except Exception as e:
            return {'success': False, 'message': f'条件删除失败: {str(e)}'}

    def conditional_filter(self, column: str, condition: str,
                           sheet_name: str = None) -> Dict[str, Any]:
        """
        筛选满足条件的行（只读，不修改文件）。

        Args:
            column:    列名或列号，None 表示扫描所有列
            condition: 条件字符串
            sheet_name: 工作表名
        Returns:
            匹配的行数据列表
        """
        try:
            ws = self.wb[sheet_name] if sheet_name else self.ws
            col_idx = self._find_column_index(column) if column else None

            # 读取表头
            headers = [cell.value for cell in ws[1]]

            matched_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                cells_to_check = [row[col_idx - 1]] if col_idx else list(row)
                if any(self._eval_condition(v, condition) for v in cells_to_check):
                    row_dict = {str(headers[i]): row[i] for i in range(len(headers)) if i < len(row)}
                    matched_rows.append(row_dict)

            return {
                'success': True,
                'message': f'筛选完成，共找到 {len(matched_rows)} 行',
                'rows': matched_rows,
                'count': len(matched_rows),
                'headers': [str(h) for h in headers if h is not None],
            }
        except Exception as e:
            return {'success': False, 'message': f'条件筛选失败: {str(e)}'}

    def close(self):
        """关闭工作簿"""
        self.wb.close()
        self.wb = None
        self.ws = None


# ─────────────────────────────────────────────────────────────────────────────
# 快捷函数
# ─────────────────────────────────────────────────────────────────────────────
def excel_get_content(file_path: str) -> Dict[str, Any]:
    """获取 Excel 内容"""
    ops = ExcelDocumentOperations(file_path)
    try:
        return ops.get_all_content()
    finally:
        ops.close()


def excel_edit_cell(file_path: str, cell_ref: str, value: Any) -> Dict[str, Any]:
    """编辑 Excel 单元格"""
    ops = ExcelDocumentOperations(file_path)
    try:
        return ops.edit_cell(cell_ref, value)
    finally:
        ops.close()


def excel_add_row(file_path: str, position: int = None, values: List[Any] = None) -> Dict[str, Any]:
    """添加 Excel 行"""
    ops = ExcelDocumentOperations(file_path)
    try:
        return ops.add_row(position, values)
    finally:
        ops.close()


def excel_extract_table(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """提取 Excel 表格"""
    ops = ExcelDocumentOperations(file_path)
    try:
        return ops.extract_table(sheet_name)
    finally:
        ops.close()


def excel_batch_fill(file_path: str, field_mapping: Dict[str, List[Any]]) -> Dict[str, Any]:
    """Excel 批量填充"""
    ops = ExcelDocumentOperations(file_path)
    try:
        return ops.batch_fill(field_mapping)
    finally:
        ops.close()

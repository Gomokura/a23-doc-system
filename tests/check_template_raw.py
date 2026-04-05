import sys
sys.path.insert(0, r'D:\桌面\a23-doc-system')
from openpyxl import load_workbook

template_path = r'D:\桌面\a23-doc-system\uploads\69eb3c13-8fdc-4e12-a461-92da0b1c5dbf.xlsx'
wb = load_workbook(template_path)
sheet = wb.active

print('模板内容:')
for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
    print(f'第{row_idx}行: {row}')
    if row_idx > 10:
        break
